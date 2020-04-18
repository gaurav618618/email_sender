import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl, smtplib, sys
wb = openpyxl.load_workbook("C:\\Users\\gaurav yadav\\Downloads\\amey.xlsx")
sheet = wb.get_sheet_by_name('Sheet1')
me = "example@email.com"   # Enter your email only works for gmail
my_password = 'password'   # Enter your password
s = smtplib.SMTP_SSL('smtp.gmail.com')
s.login(me, my_password)
message = "whats up!"
msg = MIMEMultipart('alternative')
msg['Subject'] = "Alert"
msg['From'] = me
rowe = 11
recv_email = sheet.cell(row=rowe, column=4).value
while recv_email != None:
    print(recv_email)
    msg['To'] = recv_email
    html = '<html><body><p>' + message  +'</p></body></html>'
    part2 = MIMEText(html, 'html')
    msg.attach(part2)
    s = smtplib.SMTP_SSL('smtp.gmail.com')
    s.login(me, my_password)

    s.sendmail(me, recv_email, msg.as_string())
    rowe += 1
    recv_email = sheet.cell(row=rowe, column=4).value
print('--------------------Finished sending mail------------------------------')
s.quit()